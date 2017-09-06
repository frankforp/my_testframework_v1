#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 8/16/2017 10:55 AM
# @Author  : Winnichen
# @File    : excel_reader.py
from openpyxl import Workbook
from openpyxl import load_workbook
from config import settings

class Common_ExcelReader(object):
    def __init__(self,sheet):
        self.wb=load_workbook(settings.testcase_path)
        self.ws=self.wb.get_sheet_by_name(sheet)

    def getCellValue(self,row,column):
        return self.ws.cell(row=row,column=column).value